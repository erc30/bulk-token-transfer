# Things this program needs
# pip3 install openpyxl
# pip3 install configparser
# pip3 install web3

# How to run this script
# python3 send_tokens.py

# Import config parser and web3 and spreadsheet parser 
import sys
import time
import json
import pickle
import configparser
from web3 import Web3
from datetime import datetime
from openpyxl import load_workbook

# Helper functions
def writeData(_file_name, _data_to_write):
    with open(_file_name, 'wb') as f:
        pickle.dump(_data_to_write, f)

def readData(_file_name):
    with open(_file_name, 'rb') as f:
        return pickle.load(f)

# Read the config
config = configparser.ConfigParser()
config.read('config.ini')
sendersPrivateKey = config['sender']['privatekey']
sendersPublicKey = config['sender']['publickey']
blockchainRpcEndpoint = config['blockchain']['rpc']
blockchainChainId = int(config['blockchain']['chainid'])
inputFile = config['data']['inputfile']
gasPrice = int(config['blockchain']['gasprice'])
gasLimit = int(config['blockchain']['gaslimit'])

# Connect to blockchain
w3 = Web3(Web3.HTTPProvider(blockchainRpcEndpoint))
if(w3.isConnected()):
    print(blockchainRpcEndpoint)
    print(blockchainChainId)
else:
    print('Unable to connect to {0}'.format(blockchainRpcEndpoint))

# Open the spreadsheet
workbook = load_workbook(filename = inputFile, data_only=True)
worksheet = workbook.active

# Create blank report object
outerReportObject = []

# Create unique file name for this run
timeOfRunning = datetime.today().strftime('%Y-%m-%d-%H-%M-%S')

writeData(timeOfRunning, outerReportObject)

def confirmTransactions(_list_of_transaction_hashes):
    print("Confirming a set of transactions, please wait ...")
    for i in _list_of_transaction_hashes:
        w3.eth.wait_for_transaction_receipt(i)
    print("Transactions confirmed, continuing now ...")

# Optional command line argument
startFrom = False
n = len(sys.argv)
if n == 2:
    startFromAddress = sys.argv[1]
    if(startFromAddress.startswith('0x')):
        if w3.isAddress(startFromAddress):
            startFrom = True
            print("Starting from {0}".format(startFromAddress))
else:
    print("Starting from the beginning of xlsx file")

# Iterate through the spreadsheet
depupedDict = {}
for i in range(1, worksheet.max_row+1):
    recipientAddress_temp = worksheet.cell(row=i, column=2).value
    if isinstance(recipientAddress_temp, str):
        if(recipientAddress_temp.startswith('0x')):
            try:
                recipientAddress = w3.toChecksumAddress(recipientAddress_temp)
                if w3.isAddress(recipientAddress):
                #    print("Processing address: {0}".format(recipientAddress))
                #    print("Processing amount: {0}".format(worksheet.cell(row=i, column=3).value))
                    raw_amount = worksheet.cell(row=i, column=3).value
                    #print(amount_temp)
                    amount = int(float(raw_amount))
                    print("Processing amount: {0}".format(amount))
                    if recipientAddress in depupedDict:
                        print("We alread have that address: {0}".format(recipientAddress))
                        print("Adding {0} and {1}".format(depupedDict[recipientAddress], amount))
                        temp_amount = depupedDict[recipientAddress] + amount
                        depupedDict[recipientAddress] = temp_amount
                    else:
                        depupedDict[recipientAddress] = amount
            except:
                print('Invalid address: {0}'.format(recipientAddress))
print("***** Deduped list of recipient accounts: {0} ******".format(len(depupedDict)))
next_nonce_should_be = w3.eth.get_transaction_count(sendersPublicKey,"pending") + 1
processingCommenced = False
# Iterate through the dict
tempListOfTransactionHashes = []
for key, value in depupedDict.items():
    print('startFrom: {0}, processingCommenced{1}'.format(startFrom, processingCommenced))
    if startFrom == True and processingCommenced == False:
        print('{0}{0}'.format(key, startFromAddress))
        if key == startFromAddress:
            processingCommenced = True
        else:
            continue
    else:
        processingCommenced = True
    print("Processing: {0} {1}".format(key, value))
    if value >= 1000000000000000000 and value <= 10000000000000000000000:
        if len(tempListOfTransactionHashes) == 100:
            confirmTransactions(tempListOfTransactionHashes)
            tempListOfTransactionHashes = []
        nonce = w3.eth.get_transaction_count(sendersPublicKey,"pending")
        print('Nonce: {0}'.format(nonce))
        if nonce < next_nonce_should_be:
            print("Waiting for nonce value to catchup")
            time.sleep(15)
            nonce = w3.eth.get_transaction_count(sendersPublicKey,"pending")
        print("Using nonce: {0}".format(nonce))
        transactionObject = {"nonce": nonce, "gas": gasLimit, "gasPrice": gasPrice, "from": sendersPublicKey, "to": key, "value": value, "chainId": blockchainChainId}
        next_nonce_should_be = nonce + 1
        print("Transaction object: {0}".format(transactionObject))
        signedTransaction = w3.eth.account.sign_transaction(transactionObject, sendersPrivateKey)
        #print('Signed transaction: {0}'.format(signedTransaction))
        sentTransaction = w3.eth.send_raw_transaction(signedTransaction.rawTransaction)
        transactionHash = w3.toHex(sentTransaction)
        tempListOfTransactionHashes.append(transactionHash)
        tempOuterReportObject = readData(timeOfRunning)
        tempInnerReportObject = [key, value, transactionHash]
        tempOuterReportObject.append(tempInnerReportObject)
        writeData(timeOfRunning, tempOuterReportObject)
        time.sleep(0.25)
    else:
        print("Amount must be more than 1CMT and less than 10000CMT")
        continue

if len(tempListOfTransactionHashes) > 0:
    confirmTransactions(tempListOfTransactionHashes)
